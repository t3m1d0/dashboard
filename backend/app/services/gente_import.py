# ============================================================
# app/services/gente_import.py
# Importação da Relação de Empregados - Salário Base/Líquido/Bruto
#
# FORMATO DO RELATÓRIO (MBG e demais empresas):
#   Linhas 0-4 : cabeçalho (empresa, competência, data)
#   Linha  5   : header das colunas
#   Linhas 6+  : dados dos colaboradores
#   Última linha: totalizador ("Total de Funcionários : N")
#
# COLUNAS SEM banco (9 valores):
#   Matrícula | Nome | CPF | Admissão | Cargo | Departamento |
#   Salário Base | Salário Bruto | Salário Líquido
#
# COLUNAS COM banco (12 valores):
#   Matrícula | Nome | CPF | Admissão | Cargo | Departamento |
#   Banco | Agência | Conta |
#   Salário Base | Salário Bruto | Salário Líquido
# ============================================================
import hashlib, io, uuid, re, json
from datetime import datetime, timezone
from typing import Optional
import pandas as pd
from sqlalchemy.ext.asyncio import AsyncSession
from sqlalchemy import select, and_, func

from app.models.gente import GenteFolha, GenteColaborador, GenteImportacao

MESES_MAP = {
    'janeiro':1,'fevereiro':2,'março':3,'marco':3,'abril':4,
    'maio':5,'junho':6,'julho':7,'agosto':8,'setembro':9,
    'outubro':10,'novembro':11,'dezembro':12
}
MESES_NOME = {
    1:'Janeiro',2:'Fevereiro',3:'Março',4:'Abril',5:'Maio',6:'Junho',
    7:'Julho',8:'Agosto',9:'Setembro',10:'Outubro',11:'Novembro',12:'Dezembro'
}


def _extrair_mes_arquivo(filename):
    nome = (filename or '').lower()
    for ext in ['.xlsx','.xls','.csv']: nome = nome.replace(ext,'')
    partes = re.split(r'[_ \-]', nome)
    for parte in reversed(partes):
        if parte in MESES_MAP:
            num = MESES_MAP[parte]; ano = datetime.now().year
            return num, MESES_NOME[num], f'{ano}-{num:02d}'
    for mes_nome, num in MESES_MAP.items():
        if mes_nome in nome:
            ano = datetime.now().year
            return num, MESES_NOME[num], f'{ano}-{num:02d}'
    return None, None, None


def _parse_valor(val):
    if val is None: return 0.0
    if isinstance(val, (int, float)):
        return float(val) if not (isinstance(val, float) and pd.isna(val)) else 0.0
    s = str(val).strip().replace('R$','').replace('\xa0','').replace(' ','')
    if not s or s in ('-','nan',''): return 0.0
    if ',' in s and '.' in s: return float(s.replace('.','').replace(',','.')) or 0.0
    if ',' in s: return float(s.replace(',','.')) or 0.0
    try: return float(s) or 0.0
    except: return 0.0


def _clean(val, maxlen=300):
    if val is None: return None
    s = str(val).strip()
    return None if s in ('','nan','None') else s[:maxlen]


def _parse_data(val):
    if val is None: return None
    if isinstance(val, datetime): return val.strftime('%Y-%m-%d')
    s = str(val).strip()
    if s in ('','nan','None'): return None
    return s.split(' ')[0] if ' ' in s else s


def _hash_row(data):
    fields = '|'.join(str(data.get(k,'')) for k in [
        'matricula','nome','competencia','salario_base','total_proventos','liquido'
    ])
    return hashlib.md5(fields.encode()).hexdigest()


def _read_raw(content, filename):
    fname = (filename or '').lower()

    if fname.endswith('.csv'):
        for enc in ['utf-8-sig','latin-1','utf-8','cp1252']:
            try:
                df = pd.read_csv(io.BytesIO(content), sep=None, engine='python',
                                 encoding=enc, header=None, dtype=str,
                                 encoding_errors='replace')
                return df.values.tolist()
            except: continue
        raise ValueError("Não foi possível ler o CSV.")

    # XLSX — tenta direto com openpyxl
    if fname.endswith('.xlsx'):
        try:
            from openpyxl import load_workbook
            wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
            ws = wb.active
            return [list(row) for row in ws.iter_rows(values_only=True)]
        except Exception as e:
            raise ValueError(f"Erro ao ler xlsx: {e}")

    # XLS — converte para xlsx via libreoffice (arquivo temporário)
    if fname.endswith('.xls') or fname.endswith('.xlsm'):
        import tempfile, subprocess, os
        with tempfile.TemporaryDirectory() as tmpdir:
            # Salva .xls temporário
            xls_path = os.path.join(tmpdir, 'input.xls')
            with open(xls_path, 'wb') as f:
                f.write(content)
            # Converte via libreoffice
            result = subprocess.run(
                ['libreoffice', '--headless', '--convert-to', 'xlsx',
                 '--outdir', tmpdir, xls_path],
                capture_output=True, text=True, timeout=60
            )
            xlsx_files = [f for f in os.listdir(tmpdir) if f.endswith('.xlsx')]
            if not xlsx_files:
                raise ValueError(
                    f"Não foi possível converter o arquivo .xls. "
                    f"Erro: {result.stderr[:200]}"
                )
            xlsx_path = os.path.join(tmpdir, xlsx_files[0])
            from openpyxl import load_workbook
            wb = load_workbook(xlsx_path, read_only=True, data_only=True)
            ws = wb.active
            return [list(row) for row in ws.iter_rows(values_only=True)]

    # Fallback genérico
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(content), read_only=True, data_only=True)
        ws = wb.active
        return [list(row) for row in ws.iter_rows(values_only=True)]
    except Exception as e:
        raise ValueError(f"Formato não suportado: {e}")


def _extrair_meta(rows):
    meta = {}
    for row in rows[:10]:
        for val in row:
            if val is None: continue
            s = str(val).strip()
            sl = s.lower()
            if sl.startswith('empresa:'):
                meta['empresa'] = s.split(':',1)[1].strip()
            elif sl.startswith('competência:') or sl.startswith('competencia:'):
                comp_str = s.split(':',1)[1].strip()
                partes = comp_str.split('/')
                if len(partes) == 2:
                    mes_n, ano_n = partes[0].strip().zfill(2), partes[1].strip()
                    meta['competencia_raw'] = f"{ano_n}-{mes_n}"
                    meta['mes_nome'] = MESES_NOME.get(int(mes_n), comp_str)
    return meta


def _encontrar_header(rows):
    for i, row in enumerate(rows[:15]):
        vals = [str(v).strip().lower() for v in row if v is not None]
        if any('matr' in v for v in vals) and any('nome' in v for v in vals):
            return i
    return 5


def _parse_row(row):
    vals = [v for v in row if v is not None and str(v).strip() not in ('','None','nan')]
    if len(vals) < 6: return None
    try:
        matricula = str(int(float(str(vals[0]))))
    except: return None
    if 'total' in str(vals[0]).lower(): return None
    nome = _clean(vals[1], 300) if len(vals) > 1 else None
    if not nome: return None
    cpf      = _clean(vals[2], 14) if len(vals) > 2 else None
    data_adm = _parse_data(vals[3]) if len(vals) > 3 else None
    cargo    = _clean(vals[4], 200) if len(vals) > 4 else None
    depto    = _clean(vals[5], 200) if len(vals) > 5 else None
    banco = agencia = conta = None
    has_bank = False
    if len(vals) >= 12:
        try: float(str(vals[9]).replace(',','.')); has_bank = True
        except: pass
    if has_bank:
        banco   = _clean(vals[6], 200)
        agencia = _clean(vals[7], 50)
        conta   = _clean(vals[8], 50)
        sal_base  = _parse_valor(vals[9])
        sal_bruto = _parse_valor(vals[10])
        sal_liq   = _parse_valor(vals[11])
    elif len(vals) >= 9:
        sal_base  = _parse_valor(vals[6])
        sal_bruto = _parse_valor(vals[7])
        sal_liq   = _parse_valor(vals[8])
    else:
        sal_base = sal_bruto = sal_liq = 0.0
    return {
        'matricula': matricula, 'nome': nome, 'cpf': cpf,
        'data_admissao': data_adm, 'cargo': cargo, 'departamento': depto,
        'banco': banco, 'agencia': agencia, 'conta': conta,
        'salario_base': sal_base, 'salario_bruto': sal_bruto, 'liquido': sal_liq,
    }


async def importar_folha(content, filename, competencia, empresa_id, db):
    rows = _read_raw(content, filename)
    if not rows: raise ValueError("Arquivo vazio.")
    meta = _extrair_meta(rows)
    empresa_nome = meta.get('empresa','')
    if not competencia:
        competencia = meta.get('competencia_raw')
    if not competencia:
        _, _, competencia = _extrair_mes_arquivo(filename)
    if not competencia:
        raise ValueError(
            "Competência não detectada. Informe manualmente (YYYY-MM) "
            "ou o arquivo deve conter 'Competência: MM/YYYY' no cabeçalho."
        )
    try:
        mes_int = int(competencia.split('-')[1])
        mes_nome_str = MESES_NOME.get(mes_int, competencia)
    except:
        mes_nome_str = meta.get('mes_nome', competencia)

    header_idx = _encontrar_header(rows)
    data_rows  = rows[header_idx + 1:]

    existing_result = await db.execute(
        select(GenteFolha.matricula, GenteFolha.dados_hash)
        .where(and_(GenteFolha.empresa_id == empresa_id, GenteFolha.competencia == competencia))
    )
    existing = {r.matricula: r.dados_hash for r in existing_result if r.matricula}

    total = 0; inserted = updated = skipped = errors = 0; erros_msg = []; to_insert = []

    for idx, raw_row in enumerate(data_rows):
        try:
            parsed = _parse_row(raw_row)
            if not parsed: continue
            total += 1
            data = {
                'empresa_id': empresa_id, 'competencia': competencia,
                'mes_nome': mes_nome_str, 'matricula': parsed['matricula'],
                'cpf': parsed['cpf'], 'nome': parsed['nome'],
                'empresa': empresa_nome or None, 'departamento': parsed['departamento'],
                'cargo': parsed['cargo'], 'filial': empresa_nome or None,
                'data_admissao': parsed['data_admissao'],
                'situacao': 'Ativo', 'tipo_contrato': 'CLT',
                'salario_base': parsed['salario_base'],
                'total_proventos': parsed['salario_bruto'], 'liquido': parsed['liquido'],
                'total_descontos': round(max(0, parsed['salario_bruto'] - parsed['liquido']), 2),
                'valor': parsed['salario_base'],
                'verba_codigo': None, 'verba_nome': None, 'verba_tipo': None,
                'referencia': None, 'fgts': None, 'inss': None, 'irrf': None,
                'centro_custo': None,
            }
            data['dados_hash'] = _hash_row(data)
            mat = data['matricula']
            if mat in existing:
                if existing[mat] != data['dados_hash']:
                    res = await db.execute(select(GenteFolha).where(and_(
                        GenteFolha.empresa_id == empresa_id,
                        GenteFolha.competencia == competencia,
                        GenteFolha.matricula == mat,
                    )))
                    obj = res.scalar_one_or_none()
                    if obj:
                        for k, v in data.items(): setattr(obj, k, v)
                        obj.atualizado_em = datetime.now(timezone.utc); updated += 1
                    else:
                        data['id'] = uuid.uuid4()
                        data['importado_em'] = data['atualizado_em'] = datetime.now(timezone.utc)
                        to_insert.append(data); inserted += 1
                else: skipped += 1
            else:
                data['id'] = uuid.uuid4()
                data['importado_em'] = data['atualizado_em'] = datetime.now(timezone.utc)
                to_insert.append(data); inserted += 1
            if len(to_insert) >= 300:
                db.add_all([GenteFolha(**r) for r in to_insert]); await db.flush(); to_insert = []
        except Exception as e:
            errors += 1; erros_msg.append(f"Linha {header_idx+idx+2}: {str(e)[:120]}")

    if to_insert: db.add_all([GenteFolha(**r) for r in to_insert])
    await _upsert_colaboradores(db, empresa_id, competencia, empresa_nome)
    db.add(GenteImportacao(
        empresa_id=empresa_id, tipo='folha', competencia=competencia,
        mes_nome=mes_nome_str, nome_arquivo=filename, total_linhas=total,
        inseridos=inserted, atualizados=updated, ignorados=skipped, erros=errors,
        colunas_detectadas=json.dumps(['matricula','nome','cpf','admissao','cargo','departamento','salario_base','salario_bruto','liquido'], ensure_ascii=False),
    ))
    await db.flush()
    return {
        'competencia': competencia, 'mes_nome': mes_nome_str, 'empresa': empresa_nome,
        'total_arquivo': total, 'inseridos': inserted, 'atualizados': updated,
        'ignorados': skipped, 'erros': errors, 'primeiros_erros': erros_msg[:5],
    }


async def _upsert_colaboradores(db, empresa_id, competencia, empresa_nome):
    result = await db.execute(select(GenteFolha).where(and_(
        GenteFolha.empresa_id == empresa_id, GenteFolha.competencia == competencia,
    )))
    for r in result.scalars().all():
        if not r.matricula: continue
        existing = (await db.execute(select(GenteColaborador).where(and_(
            GenteColaborador.empresa_id == empresa_id, GenteColaborador.matricula == r.matricula,
        )))).scalar_one_or_none()
        if existing:
            existing.nome = r.nome or existing.nome
            existing.cargo = r.cargo or existing.cargo
            existing.departamento = r.departamento or existing.departamento
            existing.salario_base = r.salario_base or existing.salario_base
            existing.data_admissao = r.data_admissao or existing.data_admissao
            existing.atualizado_em = datetime.now(timezone.utc)
        else:
            db.add(GenteColaborador(
                empresa_id=empresa_id, matricula=r.matricula, cpf=r.cpf,
                nome=r.nome, empresa=empresa_nome or r.empresa,
                departamento=r.departamento, cargo=r.cargo,
                filial=empresa_nome or r.filial, situacao='Ativo',
                tipo_contrato='CLT', data_admissao=r.data_admissao,
                salario_base=r.salario_base or 0,
            ))


async def get_stats(db, empresa_id, competencia=None, departamento=None, cargo=None, filial=None):
    from sqlalchemy import and_, desc
    filters = []
    if empresa_id:   filters.append(GenteFolha.empresa_id == empresa_id)
    if competencia:  filters.append(GenteFolha.competencia == competencia)
    if departamento: filters.append(GenteFolha.departamento == departamento)
    if cargo:        filters.append(GenteFolha.cargo == cargo)
    if filial:       filters.append(GenteFolha.empresa == filial)
    base = and_(*filters) if filters else True

    kpi = (await db.execute(select(
        func.count(GenteFolha.matricula.distinct()).label('total_colab'),
        func.sum(GenteFolha.salario_base).label('massa_salarial'),
        func.sum(GenteFolha.total_proventos).label('total_bruto'),
        func.sum(GenteFolha.total_descontos).label('total_descontos'),
        func.sum(GenteFolha.liquido).label('total_liquido'),
        func.avg(GenteFolha.salario_base).label('media_salario'),
        func.avg(GenteFolha.liquido).label('media_liquido'),
    ).where(base))).one()

    por_depto = (await db.execute(select(
        GenteFolha.departamento,
        func.count(GenteFolha.matricula.distinct()).label('colab'),
        func.sum(GenteFolha.salario_base).label('massa'),
        func.sum(GenteFolha.total_proventos).label('bruto'),
        func.sum(GenteFolha.liquido).label('liquido'),
        func.avg(GenteFolha.salario_base).label('media'),
    ).where(base).group_by(GenteFolha.departamento)
    .order_by(desc(func.count(GenteFolha.matricula.distinct()))))).fetchall()

    por_cargo = (await db.execute(select(
        GenteFolha.cargo,
        func.count(GenteFolha.matricula.distinct()).label('colab'),
        func.sum(GenteFolha.salario_base).label('massa'),
        func.avg(GenteFolha.salario_base).label('media'),
    ).where(base).group_by(GenteFolha.cargo)
    .order_by(desc(func.count(GenteFolha.matricula.distinct()))).limit(15))).fetchall()

    competencias = (await db.execute(select(
        GenteFolha.competencia, GenteFolha.mes_nome,
        func.count(GenteFolha.matricula.distinct()).label('colab'),
        func.sum(GenteFolha.salario_base).label('massa'),
        func.sum(GenteFolha.liquido).label('liquido'),
        func.sum(GenteFolha.total_proventos).label('bruto'),
    ).where(GenteFolha.empresa_id == empresa_id)
    .group_by(GenteFolha.competencia, GenteFolha.mes_nome)
    .order_by(GenteFolha.competencia.desc()))).fetchall()

    deptos_list   = [r[0] for r in (await db.execute(select(GenteFolha.departamento).where(base).distinct().order_by(GenteFolha.departamento))) if r[0]]
    cargos_list   = [r[0] for r in (await db.execute(select(GenteFolha.cargo).where(base).distinct().order_by(GenteFolha.cargo))) if r[0]]
    empresas_list = [r[0] for r in (await db.execute(select(GenteFolha.empresa).where(GenteFolha.empresa_id == empresa_id).distinct().order_by(GenteFolha.empresa))) if r[0]]

    return {
        'kpis': {
            'total_colaboradores': int(kpi.total_colab or 0),
            'massa_salarial':  round(float(kpi.massa_salarial or 0), 2),
            'total_bruto':     round(float(kpi.total_bruto or 0), 2),
            'total_descontos': round(float(kpi.total_descontos or 0), 2),
            'total_liquido':   round(float(kpi.total_liquido or 0), 2),
            'media_salario':   round(float(kpi.media_salario or 0), 2),
            'media_liquido':   round(float(kpi.media_liquido or 0), 2),
        },
        'por_departamento': [
            {'nome': r.departamento or 'N/A', 'colab': int(r.colab or 0),
             'massa': round(float(r.massa or 0), 2), 'bruto': round(float(r.bruto or 0), 2),
             'liquido': round(float(r.liquido or 0), 2), 'media': round(float(r.media or 0), 2)}
            for r in por_depto
        ],
        'por_cargo': [
            {'nome': r.cargo or 'N/A', 'colab': int(r.colab or 0),
             'massa': round(float(r.massa or 0), 2), 'media': round(float(r.media or 0), 2)}
            for r in por_cargo
        ],
        'competencias': [
            {'competencia': r.competencia, 'mes_nome': r.mes_nome,
             'colab': int(r.colab or 0), 'massa': round(float(r.massa or 0), 2),
             'liquido': round(float(r.liquido or 0), 2), 'bruto': round(float(r.bruto or 0), 2)}
            for r in competencias
        ],
        'filtros': {
            'departamentos': deptos_list,
            'cargos': cargos_list,
            'empresas': empresas_list,
        }
    }
